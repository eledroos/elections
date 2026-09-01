"""Read the turnout workbook that the Boston Election Department sends.

Each workbook holds two sheets with the same shape: a grid of wards down the
side and precinct numbers across the top. The first sheet holds ballot counts.
The second holds the same figures as a share of registered voters. Row 3 holds
the precinct numbers, and the last column holds the department's own ward
total.

The department does not send the number of registered voters. Because it sends
both a count and a percentage, that number divides out exactly, so
``derive_registration`` recovers it.
"""
from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path

import openpyxl

HEADER_ROW = 3
FIRST_WARD_ROW = 4
FIRST_PRECINCT_COL = 2


def key(ward: int, precinct: int) -> str:
    """Return the identifier shared by the workbook and the boundary file."""
    return f"{ward:02d}{precinct:02d}"


@dataclass
class ParsedSnapshot:
    """One workbook, read but not judged."""

    path: Path
    counts: dict[str, int] = field(default_factory=dict)
    percents: dict[str, float | None] = field(default_factory=dict)
    stated_ward_totals: dict[int, int] = field(default_factory=dict)
    stated_citywide: int = 0


def _ward_number(cell) -> int | None:
    if not isinstance(cell, str):
        return None
    parts = cell.strip().split()
    if len(parts) == 2 and parts[0].upper() == "WARD" and parts[1].isdigit():
        return int(parts[1])
    return None


def _precinct_columns(sheet) -> dict[int, int]:
    """Map a spreadsheet column to the precinct number written above it."""
    columns = {}
    for col in range(FIRST_PRECINCT_COL, sheet.max_column + 1):
        value = sheet.cell(row=HEADER_ROW, column=col).value
        if isinstance(value, int):
            columns[col] = value
    return columns


def _number(value) -> float | None:
    """Return a number, or nothing for a blank cell or a spreadsheet error."""
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        return None
    return value


def parse_workbook(path: Path, counts_sheet: int = 0, percent_sheet: int = 1) -> ParsedSnapshot:
    book = openpyxl.load_workbook(path, data_only=True)
    counts_grid = book.worksheets[counts_sheet]
    percent_grid = book.worksheets[percent_sheet]

    parsed = ParsedSnapshot(path=path)
    columns = _precinct_columns(counts_grid)
    percent_columns = _precinct_columns(percent_grid)

    for row in range(FIRST_WARD_ROW, counts_grid.max_row + 1):
        ward = _ward_number(counts_grid.cell(row=row, column=1).value)
        if ward is None:
            continue
        for col, precinct in columns.items():
            count = _number(counts_grid.cell(row=row, column=col).value)
            if count is None:
                continue
            identifier = key(ward, precinct)
            parsed.counts[identifier] = int(count)
            share = None
            for percent_col, percent_precinct in percent_columns.items():
                if percent_precinct == precinct:
                    share = _number(percent_grid.cell(row=row, column=percent_col).value)
                    break
            parsed.percents[identifier] = share
        total = _number(counts_grid.cell(row=row, column=_total_column(counts_grid)).value)
        if total is not None:
            parsed.stated_ward_totals[ward] = int(total)

    parsed.stated_citywide = sum(parsed.stated_ward_totals.values())
    return parsed


def _total_column(sheet) -> int:
    for col in range(FIRST_PRECINCT_COL, sheet.max_column + 1):
        if str(sheet.cell(row=HEADER_ROW, column=col).value).strip().upper() == "TOTAL":
            return col
    return sheet.max_column


def derive_registration(parsed: ParsedSnapshot) -> dict[str, int]:
    """Recover registered voters per precinct from the count and the share.

    A precinct with no ballots and no share gives no answer, so it is left out.
    """
    registered: dict[str, int] = {}
    for identifier, count in parsed.counts.items():
        share = parsed.percents.get(identifier)
        if not share:
            continue
        registered[identifier] = round(count / share)
    return registered
